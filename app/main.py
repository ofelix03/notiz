import logging
import os
import smtplib
import ssl
from datetime import date, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

app_dir = "/".join(__file__.split("/")[:-1])
env_file = Path(app_dir) / ".env"

load_dotenv(dotenv_path=env_file)


LOG_FILE = Path(app_dir) / "data/notiz.log"
LOG_FORMAT = "%(asctime)s:%(levelname)s %(message)s"
logging.basicConfig(filename=LOG_FILE, format=LOG_FORMAT, level=logging.INFO)


def build_email_message(name, is_due=False):
    if is_due:
        message = f"""
            <p>Hello,</p>
            
            <p>The scheduled screening {name} is happening today.</p>
            
            <p><i>This is a system generated email. Please do not response to 
            this</i></p>
        """
    else:
        message = f"""
            <p>Hello,</p>

            <p>This is a reminder that the scheduled screening {name} is 
            {reminder_days_before()} days away.</p>
            
             <p><i>This is a system generated email. Please do not response to 
            this</i></p>
        """
    return message


def reminder_days_before():
    return int(os.getenv("SEND_REMINDER_DAYS_BEFORE"))


def is_event_due(due_date):
    if not due_date:
        return False

    if isinstance(due_date, datetime):
        due_date = due_date.date()

    diff = due_date - date.today()
    return diff.days == 0


def is_event_reminder_due(due_date):
    if not due_date:
        return False

    if isinstance(due_date, datetime):
        due_date = due_date.date()

    diff = due_date - date.today()
    return diff.days == reminder_days_before()


def send_email_notification(message, subject):
    recipient_emails = os.getenv("NOTIFICATION_RECIPIENT_EMAILS")
    sender_email = os.getenv("NOTIFICATION_SENDER_EMAIL")
    msg = MIMEMultipart()
    msg["Subject"] = f"NOTIZ: {subject}"
    msg["From"] = sender_email
    msg["To"] = recipient_emails
    msg.attach(MIMEText(message, "html"))
    with smtplib.SMTP(
        host=os.getenv("SMTP_SERVER"), port=os.getenv("SMTP_PASSWORD")
    ) as s:
        s.starttls()
        s.login(os.getenv("SMTP_USER"), os.getenv("SMTP_PASSWORD"))
        s.sendmail(msg.as_string())


def read_sheet_n_send_notifications(filepath=None):
    sent_reminders = False
    sent_due_notifications = False

    if not filepath:
        file = os.getenv("XL_FILE_PATH")
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, min_col=3):
            [day, visit_date, lower_limit, due_date, upper_limit, remark] = row

            if is_event_due(due_date.value):
                sent_due_notifications = True
                subject = f"{day.value} SCREENING IS TODAY"
                logging.info(f"[{subject}] Sending event due email notification")
                send_email_notification(
                    message=build_email_message(day.value, due_date.value),
                    subject=subject,
                )
                logging.info("[{subject}] Due email notification successfully sent")

            if is_event_reminder_due(due_date.value):
                sent_reminders = True
                subject = f"{day.value} SCREEN IS {reminder_days_before()} DAYS AWAY"
                logging.info(f"[{subject}] Sending event reminder email notification")
                send_email_notification(
                    message=build_email_message(day.value, due_date.value),
                    subject=subject,
                )
                logging.info(
                    "[{subject}] Reminder email notification successfully " "sent"
                )

    if not sent_due_notifications:
        logging.info("There are no events due today. No notifications sent out")

    if not sent_reminders:
        logging.info(
            f"No reminder emails sent ou today. There are no event due in "
            f"the "
            f"next "
            f"{reminder_days_before()} days"
        )


if __name__ == "__main__":
    read_sheet_n_send_notifications()
